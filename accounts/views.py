from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from .models import User
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import user_passes_test
from .forms import CreateUserForm, UpdateUserForm, EditMyProfileForm
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def register(request):
    form = CreateUserForm()
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Account created successfully.')
            return redirect('login')
    return render(request, 'accounts/register.html', {'form': form})

@login_required
def dashboard(request):
    return render(request, 'accounts/dashboard.html') 

def logoutUser(request):
    logout(request)
    return redirect('login')  

@login_required
def users_list(request):
    users = User.objects.all()
    return render(request, 'accounts/users.html', {'users': users})

@login_required
def addUser(request):
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = form.cleaned_data['is_staff']  # Set admin flag
            user.save()
            messages.success(request, 'User created successfully.')
            return redirect('users')
    else:
        form = CreateUserForm()
    return render(request, 'accounts/new_user.html', {'form': form})

@user_passes_test(lambda u: u.is_staff)
def editUser(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)
    form = UpdateUserForm(request.POST or None, instance=user_obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Changes successfully saved')
        return redirect('users') 
    return render(request, 'accounts/edit_user.html', {'form': form, 'user_obj': user_obj})

@user_passes_test(lambda u: u.is_staff)
def deleteUser(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)
    if request.method == "POST":
        user_obj.delete()
        return redirect('users') 
    return render(request, 'accounts/confirm_delete_user.html', {'user_obj': user_obj})

@login_required
def edit_my_profile(request):
    user = request.user
    if request.method == 'POST':
        form = EditMyProfileForm(request.POST, instance=user, user=user)  # <-- user passed here
        if form.is_valid():
            new_password = form.cleaned_data.get('new_password')
            if new_password:
                user.set_password(new_password)
            form.save()
            messages.success(request, 'Your profile has been updated.')
            if new_password:
                return redirect('login')
            return redirect('edit_profile')
    else:
        form = EditMyProfileForm(instance=user, user=user)  # <-- and here

    return render(request, 'accounts/edit_profile.html', {'form': form})

def export_users_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Current Users"

    # Define your headers
    headers = ['First Name', 'Last Name', 'Email', 'Phone Number']
    ws.append(headers)

    # Query the data
    users = User.objects.all()

    for item in users:
        ws.append([
            item.first_name,
            item.last_name,  # adjust depending on your foreign key
            item.phone_number,
            item.email,
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Users.xlsx'
    wb.save(response)
    return response

@user_passes_test(lambda u: u.is_staff)
def system_audit(request):
        
    return render(request, 'accounts/dashboard.html')





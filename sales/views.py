from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Sale, SaleItem
from inventory.models import StockItem
import json
from django.utils import timezone
from django.db.models import Sum, F
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

# Create your views here.
@csrf_exempt
@login_required
def complete_sale(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cart_items = data.get('cartItems', [])
            total = data.get('total', 0)

            if not cart_items:
                return JsonResponse({'error': 'Cart is empty'}, status=400)

            sale = Sale.objects.create(total_amount=total, cashier=request.user)


            for item in cart_items:
                product_id = item.get('id')  
                quantity = item.get('quantity')
                price = item.get('price')

                try:
                    product = StockItem.objects.get(id=product_id)
                    if product.quantity < quantity:
                        return JsonResponse({'error': f"Not enough stock for {product.product_name}"}, status=400)

                    product.quantity -= quantity
                    product.save()

                    SaleItem.objects.create(
                        sale=sale,
                        product=product,
                        quantity_sold=quantity,
                        total_amount=price * quantity
                    )

                except StockItem.DoesNotExist:
                    return JsonResponse({'error': 'Product not found'}, status=404)

            return JsonResponse({'message': 'Sale completed successfully', 'sale_id': sale.id})

        except json.JSONDecodeError as e:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

def sale_receipt(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    sale_items = sale.items.select_related('product')  # thanks to related_name='items'

    context = {
        'sale': sale,
        'items': sale_items,          
        'cashier': sale.cashier,      
        'date': sale.sold_on,
          
    }
    return render(request, 'sales/receipt.html', context)

from django.db.models import Sum, F
from django.shortcuts import render
from .models import Sale

@login_required
def sales_report(request):
    query = request.GET.get('q', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Start from SaleItem instead of Sale
    sale_items = SaleItem.objects.select_related('sale', 'product').order_by('-sale__sold_on')

    if start_date and end_date:
        sale_items = sale_items.filter(sale__sold_on__date__range=[start_date, end_date])

    if query:
        sale_items = sale_items.filter(product__product_name__icontains=query)

    total_revenue = sale_items.aggregate(total=Sum('total_amount'))['total'] or 0
    total_items = sale_items.aggregate(qty=Sum('quantity_sold'))['qty'] or 0

    return render(request, 'sales/sales_report.html', {
        'sales': sale_items,
        'total_revenue': total_revenue,
        'total_items': total_items,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
    })

def export_report_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales report"

    # Define your headers
    headers = ['Product', 'Quantity Sold', 'Selling Price', 'Total Amount', 'Date', 'Sold By']
    ws.append(headers)

    # Query the data
    sale_items = SaleItem.objects.all()

    for sale in sale_items:
        ws.append([
            sale.product.product_name,
            sale.quantity_sold,
            sale.product.selling_price,
            sale.total_amount,
            sale.sale.sold_on.strftime('%Y-%m-%d'),
            sale.sale.cashier.email,
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Sales_report.xlsx'
    wb.save(response)
    return response













from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import StockForm, SupplierForm, CategoryForm
from django.contrib.auth.decorators import login_required
from .models import StockItem, Supplier, Category
from sales.models import Sale, SaleItem
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.contrib.auth import get_user_model
User = get_user_model()
from django.core.paginator import Paginator
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q

@login_required
def index(request):
    if request.user.is_authenticated:
        return redirect('dashboard')  # or 'add-stock', etc.
    return redirect('login')

@login_required
def add_stock(request):
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Stock item added successfully!")
            return redirect('add-stock')  # or wherever you want
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StockForm()
    
    return render(request, 'inventory/add_stock.html', {'form': form})

@login_required
def show_stock(request):
    query = request.GET.get('q', '')  # Default to empty string
    stockitems = StockItem.objects.all().order_by('-added_on')  # adjust as needed
    if query:
        stockitems = stockitems.filter(
            Q(product_name__icontains=query) |
            Q(supplier__name__icontains=query)
            
        )
    paginator = Paginator(stockitems, 50)  # Show 50 stocks per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'inventory/all_stock.html', {
        'page_obj': page_obj,
        'query': query,
    })

@user_passes_test(lambda u: u.is_staff)
def edit_stock(request, stockitem_id):
    stockitem_obj = get_object_or_404(StockItem, pk=stockitem_id)
    form = StockForm(request.POST or None, instance=stockitem_obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Changes successfully saved')
        return redirect('all-stock') 
    return render(request, 'inventory/edit_stock.html', {'form': form, 'stockitem_obj': stockitem_obj})

@user_passes_test(lambda u: u.is_staff)
def delete_stock(request, stockitem_id):
    stockitem_obj = get_object_or_404(StockItem, pk=stockitem_id)
    if request.method == "POST":
        stockitem_obj.delete()
        return redirect('all-stock') 
    return render(request, 'inventory/confirm_delete_stock.html', {'stockitem_obj': stockitem_obj})

@login_required
def all_suppliers(request):
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/suppliers.html', {'suppliers': suppliers})

@login_required
def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier added successfully!")
            return redirect('suppliers')  # or wherever you want
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = SupplierForm()
    
    return render(request, 'inventory/add_supplier.html', {'form': form})

@user_passes_test(lambda u: u.is_staff)
def edit_supplier(request, supplier_id):
    supplier_obj = get_object_or_404(Supplier, pk=supplier_id)
    form = SupplierForm(request.POST or None, instance=supplier_obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Changes successfully saved')
        return redirect('suppliers') 
    return render(request, 'inventory/edit_supplier.html', {'form': form, 'supplier_obj': supplier_obj})

@user_passes_test(lambda u: u.is_staff)
def delete_supplier(request, supplier_id):
    supplier_obj = get_object_or_404(Supplier, pk=supplier_id)
    if request.method == "POST":
        supplier_obj.delete()
        return redirect('suppliers') 
    return render(request, 'inventory/confirm_delete_supplier.html', {'supplier_obj': supplier_obj})

@csrf_exempt
def pos_search(request):
    query = request.GET.get('q', '')
    results = []
    if query:
        items = StockItem.objects.filter(product_name__icontains=query)
        for item in items:
            results.append({
                'id': item.id,
                'product_name': item.product_name,
                'selling_price': str(item.selling_price)
            })
    return JsonResponse({'results': results})
def pos_view(request):
    products = StockItem.objects.all()
    return render(request, 'your_pos_template.html', {'products': products})

@login_required
def stock_categories(request):
    category = Category.objects.all()
    return render(request, 'inventory/stock_category.html', {'category': category})

@login_required
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added successfully!")
            return redirect('categories')  # or wherever you want
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CategoryForm()
    return render(request, 'inventory/add_category.html', {'form': form})

@user_passes_test(lambda u: u.is_staff)
def edit_category(request, category_id):
    category_obj = get_object_or_404(Category, pk=category_id)
    form = CategoryForm(request.POST or None, instance=category_obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Changes successfully saved')
        return redirect('categories') 
    return render(request, 'inventory/edit_category.html', {'form': form, 'category_obj': category_obj})

@user_passes_test(lambda u: u.is_staff)
def delete_category(request, category_id):
    category_obj = get_object_or_404(Category, pk=category_id)
    if request.method == "POST":
        category_obj.delete()
        return redirect('categories`') 
    return render(request, 'inventory/confirm_delete_category.html', {'category_obj': category_obj})

def export_stock_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Current Stock"

    # Define your headers
    headers = ['Product Name', 'Supplier', 'Category', 'Quantity', 'Unit Price', 'Selling Price', 'Expiry Date', 'Notes', 'Date Created']
    ws.append(headers)

    # Query the data
    stock_items = StockItem.objects.all()

    for item in stock_items:
        ws.append([
            item.product_name,
            item.supplier.name,  # adjust depending on your foreign key
            item.category.name,
            item.quantity,
            item.unit_price,
            item.selling_price,
            item.expiry_date.strftime('%Y-%m-%d'),
            item.notes,
            item.added_on.strftime('%Y-%m-%d'),
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Stock.xlsx'
    wb.save(response)
    return response

def export_supplier_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Current Suppliers"

    # Define your headers
    headers = ['Name', 'Contact Info']
    ws.append(headers)

    # Query the data
    suppliers = Supplier.objects.all()

    for item in suppliers:
        ws.append([
            item.name,
            item.contact_info,    
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Suppliers.xlsx'
    wb.save(response)
    return response

def export_category_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Categories"

    # Define your headers
    headers = ['Name']
    ws.append(headers)

    # Query the data
    category = Category.objects.all()

    for item in category:
        ws.append([
            item.name, 
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Categories.xlsx'
    wb.save(response)
    return response











